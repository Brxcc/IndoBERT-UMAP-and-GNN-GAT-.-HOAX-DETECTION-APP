"""
admin_pipeline.py — Backend routes for ML Pipeline
====================================================
Endpoints:
  POST /admin/preprocess               — Run preprocessing on dataset
  GET  /admin/preprocess/history       — List preprocessed logs
  DEL  /admin/preprocess/history/{id}  — Delete a log
  GET  /admin/preprocess/download/{id} — Download preprocessed file

  POST /admin/train-pipeline           — Start training (background)
  GET  /admin/training/status/{job_id} — Poll training status
  DEL  /admin/training/status/{job_id} — Clear job log
  GET  /admin/training/results         — All model results
  GET  /admin/training/results/{id}    — Detail (with parsed settings)
  DEL  /admin/training/results/{id}    — Delete result

  POST /admin/predict-text             — Single text inference
  POST /admin/predict-bulk             — Bulk CSV/Excel inference
  GET  /admin/testing/history          — List testing sessions
  DEL  /admin/testing/history/{id}     — Delete a session
"""

from fastapi import (
    APIRouter, Depends, HTTPException, BackgroundTasks,
    UploadFile, File
)
from fastapi.responses import FileResponse
from fastapi.concurrency import run_in_threadpool
from sqlalchemy.orm import Session
from database import get_db
import models, schemas
import pandas as pd
from typing import Optional
import json, os, uuid, copy, io, tempfile
import numpy as np
import torch

from ml_core.indosbert_umap import get_indosbert, UMAPReducer
from ml_core.training_engine import train_model, construct_graph, MODELS_DIR
from ml_core.gat_network import ContentGraphGAT
from ml_core.preprocessing_utils import preprocess_text, get_stats

router = APIRouter(prefix="/admin", tags=["Admin ML Pipeline"])


# --- Helpers ------------------------------------------------------------------
def read_dataset_file(filepath: str) -> pd.DataFrame:
    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".xlsx", ".xls"):
        return pd.read_excel(filepath, engine="openpyxl")
    for enc in ["utf-8", "utf-8-sig", "latin-1", "cp1252"]:
        try:
            return pd.read_csv(filepath, encoding=enc)
        except Exception:
            continue
    return pd.read_csv(filepath, encoding="latin-1", errors="replace")


# -----------------------------------------------------------------------------
# PREPROCESSING
# -----------------------------------------------------------------------------
@router.post("/preprocess")
def preprocess_dataset(req: schemas.ProcessRequest, db: Session = Depends(get_db)):
    dataset = db.query(models.Dataset).filter(models.Dataset.id == req.dataset_id).first()
    if not dataset:
        raise HTTPException(status_code=404, detail="Dataset not found")
    try:
        import re
        df = read_dataset_file(dataset.filepath)

        # Detect text column
        text_col = None
        for col in df.columns:
            c = str(col).lower()
            if any(k in c for k in ["text", "judul", "narasi", "berita", "konten"]):
                text_col = col
                break
        if not text_col:
            text_col = df.columns[0]

        total_tokens_before = 0
        total_tokens_after  = 0
        all_slang_matches   = []
        all_stopword_matches = []

        def process_row(row_text):
            nonlocal total_tokens_before, total_tokens_after
            tokens_before = len(re.findall(r'\b\w+\b', str(row_text)))
            total_tokens_before += tokens_before

            clean, slang_m, sw_m = preprocess_text(
                row_text,
                convert_slang=req.convert_slang,
                remove_stopwords=req.remove_stopwords,
            )
            tokens_after = len(re.findall(r'\b\w+\b', clean))
            total_tokens_after += tokens_after
            all_slang_matches.extend(slang_m)
            all_stopword_matches.extend(sw_m)
            return clean

        df[text_col] = df[text_col].astype(str).apply(process_row)
        df = df[df[text_col].str.strip() != ""]

        slang_table, stopword_table = get_stats(
            all_slang_matches, all_stopword_matches, df, text_col
        )

        clean_path = os.path.splitext(dataset.filepath)[0] + "_clean.csv"
        df.to_csv(clean_path, index=False)

        version_str = "Cleaned v1"
        if req.convert_slang:    version_str += " + Slang"
        if req.remove_stopwords: version_str += " + Stopwords"

        log = models.PreprocessedLog(
            dataset_name=dataset.name, version=version_str, filepath=clean_path
        )
        db.add(log); db.commit(); db.refresh(log)

        reduction_pct = round(
            (1 - total_tokens_after / total_tokens_before) * 100, 2
        ) if total_tokens_before > 0 else 0.0

        return {
            "message": f"Preprocessing complete: {len(df)} rows",
            "path":    clean_path,
            "rows":    len(df),
            "stats": {
                "tokens_before":    total_tokens_before,
                "tokens_after":     total_tokens_after,
                "reduction_pct":    reduction_pct,
                "slang_total":      len(all_slang_matches),
                "slang_unique":     len(set(s[0] for s in all_slang_matches)),
                "slang_list":       slang_table[:100],
                "stopword_total":   len(all_stopword_matches),
                "stopword_unique":  len(set(all_stopword_matches)),
                "stopword_list":    stopword_table[:100],
                "reduction_breakdown": {
                    "slang":     len(all_slang_matches),
                    "stopwords": len(all_stopword_matches),
                },
            },
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Preprocessing failed: {e}")


@router.get("/preprocess/history")
def get_preprocess_history(db: Session = Depends(get_db)):
    logs = db.query(models.PreprocessedLog).order_by(models.PreprocessedLog.timestamp.desc()).all()
    return [
        {
            "id": l.id, "dataset_name": l.dataset_name, "version": l.version,
            "filepath": l.filepath,
            "timestamp": l.timestamp.isoformat() if l.timestamp else None,
        }
        for l in logs
    ]


@router.delete("/preprocess/history/{log_id}")
def delete_preprocess_history(log_id: int, db: Session = Depends(get_db)):
    log = db.query(models.PreprocessedLog).filter(models.PreprocessedLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(log); db.commit()
    return {"message": "Deleted"}


@router.get("/preprocess/download/{log_id}")
def download_preprocessed(log_id: int, db: Session = Depends(get_db)):
    """Unduh file dataset yang sudah diproses (bukan dataset mentah)."""
    log = db.query(models.PreprocessedLog).filter(models.PreprocessedLog.id == log_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Log not found")
    if not log.filepath or not os.path.exists(log.filepath):
        raise HTTPException(status_code=404, detail="File not found on server")
    filename = os.path.basename(log.filepath)
    return FileResponse(log.filepath, media_type="text/csv", filename=filename)


# -----------------------------------------------------------------------------
# TRAINING PIPELINE
# -----------------------------------------------------------------------------
training_jobs: dict = {}


def background_train_task(job_id: str, dataset_id: int, settings: dict):
    from database import SessionLocal
    db = SessionLocal()
    try:
        dataset = db.query(models.Dataset).filter(models.Dataset.id == dataset_id).first()
        if not dataset:
            training_jobs[job_id]["status"] = "error"
            training_jobs[job_id]["error_msg"] = "Dataset not found"
            return

        # -- Load & Parse Dataset --------------------------------------
        training_jobs[job_id]["status"] = "loading_data"
        df = read_dataset_file(dataset.filepath)
        df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")

        # Auto-detect header row
        for idx, row in df.head(10).iterrows():
            row_str = " ".join(str(x).lower() for x in row.values)
            if any(k in row_str for k in ["judul", "text", "narasi", "berita"]) and \
               any(k in row_str for k in ["label", "target", "kelas"]):
                df.columns = df.loc[idx]
                df = df.drop(index=df.index[:idx+1]).reset_index(drop=True)
                break

        # Auto-detect text/label column
        text_col = label_col = None
        for col in df.columns:
            c = str(col).lower()
            if not text_col  and any(k in c for k in ["text", "judul", "narasi", "berita", "konten"]): text_col  = col
            if not label_col and any(k in c for k in ["label", "target", "kelas", "kategori"]):          label_col = col

        if not text_col:
            str_cols = [c for c in df.columns if df[c].dtype == object]
            text_col = max(str_cols, key=lambda c: df[c].str.len().mean()) if str_cols else df.columns[0]
        if not label_col:
            label_col = next((c for c in df.columns if c != text_col and df[c].nunique() in [2,3,4]), df.columns[-1])

        df = df.dropna(subset=[text_col, label_col])
        texts      = [str(x) for x in df[text_col].tolist()]
        raw_labels = df[label_col].tolist()
        unique_vals = sorted(set(raw_labels))

        labels = [unique_vals.index(v) for v in raw_labels] if any(isinstance(x, str) for x in raw_labels) else \
                 [0 if v == min(unique_vals) else 1 for v in raw_labels] if len(unique_vals) == 2 else \
                 [int(v) for v in raw_labels]

        texts = texts[:len(labels)]; labels = labels[:len(texts)]
        num_classes = len(unique_vals)
        training_jobs[job_id]["label_map"] = {str(i): str(v) for i, v in enumerate(unique_vals)}

        # -- IndoSIndoSBERT Sentence Embedding ----------------------------------------
        training_jobs[job_id]["status"] = "extracting_features"
        model_name   = settings.get("indosbert_model_name", "firqaaa/indo-sentence-bert-base")
        max_seq_len  = settings.get("max_seq_length", 128)
        batch_size   = settings.get("indo_batch_size", 16)

        # Set global seed before embedding extraction (uses transformers.set_seed too)
        random_seed = int(settings.get("random_seed", 42))
        from ml_core.training_engine import set_global_seed
        set_global_seed(random_seed)

        extractor    = get_indosbert(model_name)
        embeddings   = extractor.get_embeddings(texts, batch_size=batch_size, max_length=max_seq_len)

        # -- UMAP -----------------------------------------------------
        algo_mode = settings.get("algorithm_mode", "hybrid")
        use_umap  = settings.get("use_umap", True) and algo_mode != "indosbert_only"

        if use_umap:
            training_jobs[job_id]["status"] = "reducing_dimensions"
            emb_np       = np.array(embeddings, dtype=np.float32)
            n_comp       = min(settings.get("umap_n_components", 64), emb_np.shape[1])
            n_neighbors  = min(settings.get("umap_n_neighbors",  15), len(emb_np) - 1)
            reducer      = UMAPReducer(
                n_components=n_comp,
                n_neighbors=n_neighbors,
                min_dist=settings.get("umap_min_dist",    0.1),
                metric=settings.get("umap_metric",        "cosine"),
                random_state=settings.get("umap_random_state", 42),
            )
            embeddings = reducer.fit_transform(emb_np)
        else:
            embeddings = np.array(embeddings, dtype=np.float32)

        # -- GAT Training ----------------------------------------------
        use_gat = settings.get("use_gat", True) and algo_mode == "hybrid"

        if use_gat:
            training_jobs[job_id]["status"] = "training_gnn"
            logs, final_model, saved_path, best_metrics = train_model(
                embeddings, labels, settings,
                job_id=job_id, training_jobs=training_jobs,
                num_classes=num_classes,
                umap_reducer=reducer if use_umap else None,
            )
        else:
            # IndoSBERT-only or IndoSBERT+UMAP — simple logistic evaluation
            training_jobs[job_id]["status"] = "training_gnn"
            from sklearn.linear_model import LogisticRegression
            from sklearn.model_selection import cross_val_predict
            from ml_core.training_engine import compute_metrics
            clf    = LogisticRegression(max_iter=1000, class_weight="balanced")
            preds  = cross_val_predict(clf, embeddings, labels, cv=min(5, len(set(labels))))
            metrics = compute_metrics(np.array(labels), preds)
            logs   = [{"iterasi": "F1", "epoch": 1, "fold": 1, **metrics, "loss": 0.0, "token_info": "LogReg CV", "is_best": True}]
            saved_path   = None
            best_metrics = metrics
            if job_id and training_jobs and job_id in training_jobs:
                training_jobs[job_id]["logs"]         = logs
                training_jobs[job_id]["best_metrics"] = best_metrics
            final_model = None

        training_jobs[job_id]["status"] = "completed"

        # -- Save to DB ------------------------------------------------
        final_log = best_metrics or {}
        db_model  = models.ModelTrainingResult(
            model_name      = settings.get("model_name", "New Model"),
            dataset_id      = dataset.id,
            split_ratio     = settings.get("data_split_ratio", f"{settings.get('train_ratio',80)}/{settings.get('test_ratio',20)}"),
            accuracy        = final_log.get("akurasi", 0),
            precision       = final_log.get("presisi", 0),
            recall          = final_log.get("recall",  0),
            f1_score        = final_log.get("f1",      0),
            mcc             = final_log.get("mcc",     0),
            macro_average   = final_log.get("macro_average", 0),
            weighted_average= final_log.get("weighted_average", 0),
            roc_auc         = final_log.get("roc_auc", 0),
            mean_std        = final_log.get("mean_std", 0),
            algorithm_mode  = algo_mode,
            settings_json   = json.dumps(settings),
            epoch_logs_json = json.dumps(logs),
            best_model_path = saved_path,
        )
        db.add(db_model); db.commit(); db.refresh(db_model)
        training_jobs[job_id]["result_id"] = db_model.id

    except Exception as e:
        training_jobs[job_id]["status"]    = "error"
        training_jobs[job_id]["error_msg"] = str(e)
        import traceback; traceback.print_exc()
    finally:
        db.close()


@router.post("/train-pipeline")
def train_pipeline(settings: schemas.TrainingSettings, background_tasks: BackgroundTasks):
    job_id = str(uuid.uuid4())
    training_jobs[job_id] = {
        "status": "starting", "logs": [], "error_msg": None,
        "best_model_path": None, "best_metrics": {}, "result_id": None, "label_map": {},
    }
    background_tasks.add_task(background_train_task, job_id, settings.dataset_id, settings.model_dump())
    return {"message": "Training started", "job_id": job_id}


@router.get("/training/status/{job_id}")
def get_training_status(job_id: str):
    if job_id not in training_jobs:
        return {"error": "Job ID not found"}
    return training_jobs[job_id]


@router.delete("/training/status/{job_id}")
def clear_training_status(job_id: str):
    if job_id in training_jobs:
        if training_jobs[job_id]["status"] in ["completed", "error"]:
            del training_jobs[job_id]
        else:
            training_jobs[job_id]["logs"] = []
    return {"message": "Done"}


@router.get("/training/results")
def get_training_results(db: Session = Depends(get_db)):
    results = db.query(models.ModelTrainingResult).order_by(models.ModelTrainingResult.timestamp.desc()).all()
    return [_serialize_result(r, db) for r in results]


@router.get("/training/results/{result_id}")
def get_training_result_detail(result_id: int, db: Session = Depends(get_db)):
    r = db.query(models.ModelTrainingResult).filter(models.ModelTrainingResult.id == result_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Not found")
    data = _serialize_result(r, db)
    data["settings"] = json.loads(r.settings_json) if r.settings_json else {}
    data["epoch_logs"] = json.loads(r.epoch_logs_json) if r.epoch_logs_json else []
    return data


@router.delete("/training/results/{result_id}")
def delete_training_result(result_id: int, db: Session = Depends(get_db)):
    r = db.query(models.ModelTrainingResult).filter(models.ModelTrainingResult.id == result_id).first()
    if not r:
        return {"error": "Not found"}
    if r.best_model_path and os.path.exists(r.best_model_path):
        try: os.remove(r.best_model_path)
        except: pass
    db.delete(r); db.commit()
    return {"message": "Deleted"}


from pydantic import BaseModel
class RenameModelRequest(BaseModel):
    new_name: str

@router.put("/training/results/{result_id}/rename")
def rename_training_result(result_id: int, req: RenameModelRequest, db: Session = Depends(get_db)):
    r = db.query(models.ModelTrainingResult).filter(models.ModelTrainingResult.id == result_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Not found")
    r.model_name = req.new_name
    db.commit()
    return {"message": "Renamed successfully", "new_name": r.model_name}


def _serialize_result(r, db=None) -> dict:
    dataset_name = None
    if db is not None:
        try:
            ds = db.query(models.Dataset).filter(models.Dataset.id == r.dataset_id).first()
            dataset_name = ds.dataset_label or ds.name if ds else None
        except Exception:
            pass
    # Parse settings_json so frontend receives an object, not a raw string
    parsed_settings = {}
    if r.settings_json:
        try:
            parsed_settings = json.loads(r.settings_json)
        except Exception:
            parsed_settings = {}
    return {
        "id": r.id, "model_name": r.model_name, "dataset_id": r.dataset_id,
        "dataset_name": dataset_name,
        "split_ratio": r.split_ratio, "accuracy": r.accuracy, "precision": r.precision,
        "recall": r.recall, "f1_score": r.f1_score, "mcc": r.mcc,
        "macro_average": r.macro_average, "weighted_average": r.weighted_average,
        "roc_auc": r.roc_auc, "mean_std": r.mean_std,
        "algorithm_mode": r.algorithm_mode,
        "settings_json": r.settings_json,   # keep raw string for backward compat
        "settings": parsed_settings,         # parsed object for frontend use
        "best_model_path": r.best_model_path,
        "timestamp": r.timestamp.isoformat() if r.timestamp else None,
    }


# -----------------------------------------------------------------------------
# DATA SPLIT RATIO TRIAL
# -----------------------------------------------------------------------------
from pydantic import BaseModel as _BM
from typing import List as _List

class SplitTrialRequest(_BM):
    ratios: _List[str]   # e.g. ["70/30", "80/20", "90/10"]
    epoch: int = 3
    learning_rate: float = 2e-5
    dataset_id: int = 0


@router.post("/split-trial")
def run_split_trial(req: SplitTrialRequest):
    """
    Lightweight split-ratio trial simulation.
    Returns mock evaluation metrics per ratio so the UI can display
    a comparison table and pick the best ratio automatically.
    For a real trial, this would swap in actual IndoSBERT fine-tuning.
    """
    import random, math
    random.seed(42)
    results = []
    for ratio_str in req.ratios:
        try:
            train_pct = int(ratio_str.split("/")[0])
        except Exception:
            train_pct = 80

        # Quality factor: ratios near 70-80% tend to be better
        quality = 1.0 - abs(train_pct - 75) / 100.0
        base = 0.70 + quality * 0.20 + random.uniform(-0.03, 0.03)
        base = max(0.55, min(0.98, base))

        per_epoch = []
        for ep in range(1, req.epoch + 1):
            factor = 1.0 - math.exp(-ep * 0.8)
            acc  = min(0.99, base * factor + random.uniform(-0.01, 0.01))
            f1   = min(0.99, acc - random.uniform(0.00, 0.03))
            prec = min(0.99, f1 + random.uniform(-0.02, 0.02))
            rec  = min(0.99, f1 + random.uniform(-0.02, 0.02))
            loss = max(0.05, 1.5 * math.exp(-ep * 0.6) + random.uniform(-0.05, 0.05))
            per_epoch.append({
                "epoch": ep,
                "accuracy": round(acc, 4),
                "f1_score": round(f1, 4),
                "precision": round(prec, 4),
                "recall": round(rec, 4),
                "loss": round(loss, 4),
            })

        best_ep = max(per_epoch, key=lambda x: x["f1_score"])
        results.append({
            "ratio": ratio_str,
            "best_epoch": best_ep["epoch"],
            "accuracy": best_ep["accuracy"],
            "f1_score": best_ep["f1_score"],
            "precision": best_ep["precision"],
            "recall": best_ep["recall"],
            "loss": best_ep["loss"],
            "per_epoch": per_epoch,
        })

    # Mark best ratio
    best = max(results, key=lambda x: x["f1_score"])
    for r in results:
        r["is_best"] = (r["ratio"] == best["ratio"])

    return {"results": results, "best_ratio": best["ratio"]}


# -----------------------------------------------------------------------------
# INFERENCE — Single Text
# -----------------------------------------------------------------------------
def _load_model_and_predict(texts: list, model_record, db) -> list:
    """
    Helper: load model, embed texts, predict.
    Returns list of {predicted_label, predicted_class, confidence, probabilities}.
    """
    if not model_record.best_model_path or not os.path.exists(model_record.best_model_path):
        raise HTTPException(
            status_code=404,
            detail=f"Model file '{model_record.model_name}' not found. Please retrain the model first."
        )

    device     = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    checkpoint = torch.load(model_record.best_model_path, map_location=device, weights_only=False)

    gat_model = ContentGraphGAT(
        in_channels    = checkpoint["in_channels"],
        hidden_channels= checkpoint["hidden_channels"],
        out_channels   = checkpoint["out_channels"],
        heads          = checkpoint["heads"],
        dropout        = checkpoint["dropout"],
        num_layers     = checkpoint.get("num_layers", 2),
    ).to(device)
    gat_model.load_state_dict(checkpoint["model_state_dict"])
    gat_model.eval()

    settings   = checkpoint.get("settings", {})
    model_name = settings.get("indosbert_model_name", "firqaaa/indo-sentence-bert-base")
    max_len    = settings.get("max_seq_length", 128)
    batch_sz   = settings.get("indo_batch_size", 16)
    extractor  = get_indosbert(model_name)
    embeddings = extractor.get_embeddings(texts, batch_size=batch_sz, max_length=max_len)
    emb_np     = np.array(embeddings, dtype=np.float32)

    # -- Apply saved UMAP (WAJIB jika dipakai saat training) --
    umap_reducer = checkpoint.get("umap_reducer")
    if umap_reducer is not None:
        emb_np = umap_reducer.transform(emb_np)

    # -- Apply saved StandardScaler normalisasi (WAJIB: harus sama saat training) --
    scaler_mean  = checkpoint.get("scaler_mean")
    scaler_scale = checkpoint.get("scaler_scale")
    if scaler_mean is not None and scaler_scale is not None:
        mean_arr  = np.array(scaler_mean,  dtype=np.float32)
        scale_arr = np.array(scaler_scale, dtype=np.float32)
        # Pastikan dimensi cocok sebelum normalisasi
        min_dim  = min(emb_np.shape[1], len(mean_arr))
        emb_np   = (emb_np[:, :min_dim] - mean_arr[:min_dim]) / (scale_arr[:min_dim] + 1e-8)
        emb_np   = emb_np.astype(np.float32)

    # Dimension alignment (setelah normalisasi)
    in_ch = checkpoint["in_channels"]
    if emb_np.shape[1] != in_ch:
        if emb_np.shape[1] > in_ch:
            emb_np = emb_np[:, :in_ch]
        else:
            pad    = np.zeros((emb_np.shape[0], in_ch - emb_np.shape[1]), dtype=np.float32)
            emb_np = np.concatenate([emb_np, pad], axis=1)

    n  = len(texts)
    x  = torch.tensor(emb_np, dtype=torch.float).to(device)
    if n == 1:
        edge_index = torch.tensor([[0], [0]], dtype=torch.long).to(device)
    else:
        from ml_core.training_engine import construct_graph
        edge_index = construct_graph(emb_np, k_neighbors=min(10, n-1)).to(device)

    proba_all = gat_model.predict_proba(x, edge_index).cpu()   # (N, C)
    label_map = ["Fakta", "Hoaks"]
    out_channels = checkpoint["out_channels"]

    results = []
    for i in range(n):
        prob       = proba_all[i].tolist()
        pred_cls   = int(torch.argmax(proba_all[i]).item())
        pred_label = label_map[pred_cls] if pred_cls < len(label_map) else f"Kelas {pred_cls}"
        confidence = round(max(prob) * 100, 2)
        results.append({
            "predicted_label": pred_label,
            "predicted_class": pred_cls,
            "confidence":      confidence,
            "probabilities":   [round(p, 4) for p in prob],
        })
    return results


@router.post("/predict-text", response_model=schemas.PredictTextResponse)
def predict_text(req: schemas.PredictTextRequest, db: Session = Depends(get_db)):
    if req.model_id:
        record = db.query(models.ModelTrainingResult).filter(models.ModelTrainingResult.id == req.model_id).first()
    else:
        record = db.query(models.ModelTrainingResult).filter(
            models.ModelTrainingResult.best_model_path != None
        ).order_by(models.ModelTrainingResult.timestamp.desc()).first()

    if not record:
        raise HTTPException(status_code=404, detail="No trained model found. Please train a model first.")

    try:
        preds = _load_model_and_predict([req.text], record, db)
        p     = preds[0]
        return schemas.PredictTextResponse(
            predicted_label=p["predicted_label"],
            predicted_class=p["predicted_class"],
            probabilities=p["probabilities"],
            model_name=record.model_name or "Model",
            model_accuracy=record.accuracy or 0.0,
        )
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Prediction failed: {e}")


@router.post("/predict-text/explain")
def predict_text_explain(req: schemas.PredictTextRequest, db: Session = Depends(get_db)):
    if req.model_id:
        record = db.query(models.ModelTrainingResult).filter(models.ModelTrainingResult.id == req.model_id).first()
    else:
        record = db.query(models.ModelTrainingResult).filter(
            models.ModelTrainingResult.best_model_path != None
        ).order_by(models.ModelTrainingResult.timestamp.desc()).first()

    if not record:
        raise HTTPException(status_code=404, detail="No trained model found.")

    try:
        from lime.lime_text import LimeTextExplainer
        import numpy as np

        def predict_proba_wrapper(texts_list):
            results = _load_model_and_predict(texts_list, record, db)
            probs = [res["probabilities"] for res in results]
            return np.array(probs)

        # LIME expects classes: 0 -> Fakta, 1 -> Hoaks
        explainer = LimeTextExplainer(class_names=["Fakta", "Hoaks"])
        
        # We'll use 50 samples for speed. We can adjust if needed.
        exp = explainer.explain_instance(
            req.text, 
            predict_proba_wrapper, 
            num_features=15, 
            num_samples=50
        )
        
        # as_list returns [('word', weight), ...]
        return {
            "model_name": record.model_name,
            "explanation": exp.as_list(),
            "text": req.text
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Explanation failed: {e}")


# -----------------------------------------------------------------------------
# INFERENCE — Bulk (File Upload)
# -----------------------------------------------------------------------------
@router.post("/predict-bulk")
async def predict_bulk(
    file: UploadFile = File(...),
    model_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """Prediksi massal dari file CSV/Excel yang diupload."""
    # Load model record
    if model_id:
        record = db.query(models.ModelTrainingResult).filter(models.ModelTrainingResult.id == model_id).first()
    else:
        record = db.query(models.ModelTrainingResult).filter(
            models.ModelTrainingResult.best_model_path != None
        ).order_by(models.ModelTrainingResult.timestamp.desc()).first()

    if not record:
        raise HTTPException(status_code=404, detail="No trained model found.")

    # Parse uploaded file
    content = await file.read()
    fname   = file.filename or "upload"
    ext     = os.path.splitext(fname)[1].lower()

    try:
        if ext in (".xlsx", ".xls"):
            df = pd.read_excel(io.BytesIO(content), engine="openpyxl")
        else:
            for enc in ["utf-8", "utf-8-sig", "latin-1"]:
                try:
                    df = pd.read_csv(io.BytesIO(content), encoding=enc); break
                except Exception:
                    continue
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {e}")

    if df.empty:
        raise HTTPException(status_code=400, detail="File is empty or invalid")

    # Detect text column
    text_col = None
    for col in df.columns:
        c = str(col).lower()
        if any(k in c for k in ["text", "judul", "narasi", "berita", "konten", "content", "news", "tweet"]):
            text_col = col
            break

    if not text_col:
        str_cols  = [c for c in df.columns if df[c].dtype == object]
        if str_cols:
            text_col = max(str_cols, key=lambda c: df[c].astype(str).str.len().mean())
        elif len(df.columns) > 0:
            text_col = df.columns[0]
            
    if not text_col:
        raise HTTPException(status_code=400, detail="Could not identify a valid text column.")

    texts = [str(x).strip() for x in df[text_col].fillna("").tolist()]
    texts = [t if t else "(empty)" for t in texts]

    try:
        preds = await run_in_threadpool(_load_model_and_predict, texts, record, db)
    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Prediction failed: {e}")

    predictions = [
        {
            "row_index":       i,
            "text":            texts[i][:300],
            "predicted_label": p["predicted_label"],
            "predicted_class": p["predicted_class"],
            "confidence":      p["confidence"],
            "probabilities":   p["probabilities"],
        }
        for i, p in enumerate(preds)
    ]

    # Detect label column for metric evaluation
    label_col = None
    for col in df.columns:
        c = str(col).lower()
        if any(k in c for k in ["label", "target", "kelas", "class", "hoaks_status", "status"]):
            label_col = col
            break

    metrics_res = None
    if label_col:
        try:
            from ml_core.training_engine import compute_metrics
            import numpy as np
            raw_labels = df[label_col].fillna("").astype(str).str.lower()
            y_true = []
            for rl in raw_labels:
                if any(k in rl for k in ["hoax", "hoaks", "1", "palsu"]):
                    y_true.append(1)
                else:
                    y_true.append(0)
            
            y_true_arr = np.array(y_true)
            y_pred_arr = np.array([p["predicted_class"] for p in preds])
            y_prob_arr = np.array([p["probabilities"] for p in preds])
            
            # Hitung metrik keseluruhan pada dataset test penuh
            metrics_res = compute_metrics(y_true_arr, y_pred_arr, y_prob_arr)
            
            # --- BOOTSTRAP TESTING untuk menghitung Mean Std ---
            # Jika sampel cukup besar, kita lakukan simulasi Bootstrap 5 kali (pengambilan acak 80% data)
            n_samples = len(y_true_arr)
            if n_samples > 10:
                n_iterations = 5
                f1_scores = []
                for _ in range(n_iterations):
                    # Subsampling secara acak dengan pengembalian (replace=True) sebanyak 80%
                    idx = np.random.choice(n_samples, size=int(n_samples * 0.8), replace=True)
                    try:
                        fold_res = compute_metrics(y_true_arr[idx], y_pred_arr[idx], y_prob_arr[idx])
                        f1_scores.append(fold_res["f1"])
                    except:
                        pass
                
                # Jika berhasil, simpangan baku F1-Score ini akan menjadi Mean Std kita
                if len(f1_scores) > 1:
                    metrics_res["mean_std"] = round(float(np.std(f1_scores)), 4)
                    
        except Exception as e:
            print(f"Failed to compute testing metrics: {e}")

    # Save testing history
    try:
        hist = models.TestingHistory(
            input_type  = "file",
            filename    = fname,
            model_id    = record.id,
            model_name  = record.model_name,
            total_rows  = len(texts),
            result_json = json.dumps(predictions[:500]),   # cap storage
            accuracy    = metrics_res["akurasi"] if metrics_res else None,
            precision   = metrics_res["presisi"] if metrics_res else None,
            recall      = metrics_res["recall"] if metrics_res else None,
            f1_score    = metrics_res["f1"] if metrics_res else None,
            mcc         = metrics_res["mcc"] if metrics_res else None,
            macro_average= metrics_res["macro_average"] if metrics_res else None,
            weighted_average= metrics_res["weighted_average"] if metrics_res else None,
            roc_auc     = metrics_res["roc_auc"] if metrics_res else None,
            mean_std    = metrics_res["mean_std"] if metrics_res else None,
        )
        db.add(hist); db.commit()
    except Exception:
        pass

    return {
        "total":        len(texts),
        "model_name":   record.model_name,
        "model_accuracy": record.accuracy,
        "predictions":  predictions,
    }


# -----------------------------------------------------------------------------
# TESTING HISTORY
# -----------------------------------------------------------------------------
@router.get("/testing/history")
def get_testing_history(db: Session = Depends(get_db)):
    items = db.query(models.TestingHistory).order_by(models.TestingHistory.timestamp.desc()).limit(50).all()
    return [
        {
            "id": h.id, "input_type": h.input_type, "filename": h.filename,
            "model_id": h.model_id, "model_name": h.model_name,
            "total_rows": h.total_rows,
            "accuracy": h.accuracy, "precision": h.precision, "recall": h.recall,
            "f1_score": h.f1_score, "mcc": h.mcc,
            "macro_average": h.macro_average, "weighted_average": h.weighted_average,
            "roc_auc": h.roc_auc, "mean_std": h.mean_std,
            "timestamp": h.timestamp.isoformat() if h.timestamp else None,
        }
        for h in items
    ]


@router.get("/testing/history/{hist_id}")
def get_testing_history_detail(hist_id: int, db: Session = Depends(get_db)):
    h = db.query(models.TestingHistory).filter(models.TestingHistory.id == hist_id).first()
    if not h:
        raise HTTPException(status_code=404, detail="Not found")
    
    # parse json safely
    result_data = []
    if h.result_json:
        try:
            result_data = json.loads(h.result_json)
        except:
            pass

    return {
        "id": h.id, "input_type": h.input_type, "filename": h.filename,
        "model_id": h.model_id, "model_name": h.model_name,
        "total_rows": h.total_rows,
        "accuracy": h.accuracy, "precision": h.precision, "recall": h.recall,
        "f1_score": h.f1_score, "mcc": h.mcc,
        "macro_average": h.macro_average, "weighted_average": h.weighted_average,
        "roc_auc": h.roc_auc, "mean_std": h.mean_std,
        "timestamp": h.timestamp.isoformat() if h.timestamp else None,
        "result_json": result_data
    }


@router.delete("/testing/history/{hist_id}")
def delete_testing_history(hist_id: int, db: Session = Depends(get_db)):
    h = db.query(models.TestingHistory).filter(models.TestingHistory.id == hist_id).first()
    if not h:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(h); db.commit()
    return {"message": "Deleted"}


# -----------------------------------------------------------------------------
# EXCEL EXPORT
# -----------------------------------------------------------------------------
@router.get("/export/excel/{result_id}")
def export_training_excel(result_id: int, db: Session = Depends(get_db)):
    """
    Export training results as a 5-sheet Excel file.
    Sheets: Parameters | Training Metrics | Testing Metrics | Comparison | Per-Epoch History
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime as dt
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    r = db.query(models.ModelTrainingResult).filter(models.ModelTrainingResult.id == result_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Training result not found")

    settings   = json.loads(r.settings_json)  if r.settings_json   else {}
    epoch_logs = json.loads(r.epoch_logs_json) if r.epoch_logs_json else []

    # Find matching test history
    test_hist = db.query(models.TestingHistory).filter(
        models.TestingHistory.model_id == result_id,
        models.TestingHistory.accuracy != None
    ).order_by(models.TestingHistory.timestamp.desc()).first()

    # ── Helper styles ──────────────────────────────────────────────────
    def header_style(cell, bg="#1E3A5F"):
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill(fill_type="solid", fgColor=bg.replace("#",""))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def section_style(cell, bg="#E8F0FE"):
        cell.font      = Font(bold=True, size=10)
        cell.fill      = PatternFill(fill_type="solid", fgColor=bg.replace("#",""))
        cell.alignment = Alignment(horizontal="left", vertical="center")

    def value_style(cell, center=False):
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center"
        )

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    wb = openpyxl.Workbook()

    # ── SHEET 1: Parameters ───────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Parameters"
    ws1.freeze_panes = "A2"

    PARAM_DESCRIPTIONS = {
        "model_name":          "Name of the trained model",
        "algorithm_mode":      "Pipeline mode: hybrid | indosbert_umap | indosbert_only",
        "max_seq_length":      "Maximum token sequence length for IndoSBERT",
        "indo_learning_rate":  "Learning rate for IndoSBERT fine-tuning",
        "indo_batch_size":     "Batch size for IndoSBERT embedding extraction",
        "indo_epoch":          "Number of epochs for IndoSBERT",
        "indo_fold":           "Number of Stratified K-Fold cross-validation folds",
        "weight_decay":        "L2 regularization weight decay coefficient",
        "warmup_ratio":        "Fraction of training steps used for LR warmup",
        "dropout_rate":        "Dropout rate applied to classifier head",
        "random_seed":         "Global random seed for reproducibility",
        "indosbert_hidden_dim": "Hidden dimension for custom IndoSBERT adapter layers",
        "indosbert_num_heads":  "Attention heads for custom IndoSBERT adapter layers",
        "early_stop_patience": "Epochs without improvement before early stopping",
        "use_umap":            "Whether UMAP dimensionality reduction is enabled",
        "umap_n_components":   "Output dimensions of UMAP reduction",
        "umap_n_neighbors":    "Number of neighbors for UMAP graph construction",
        "umap_min_dist":       "Minimum distance between points in UMAP embedding",
        "umap_metric":         "Distance metric for UMAP (cosine | euclidean | manhattan)",
        "use_gat":             "Whether Graph Attention Network is enabled",
        "gat_hidden_dim":      "Hidden channel size in GAT layers",
        "gat_num_heads":       "Number of attention heads in GAT",
        "gat_dropout":         "Dropout rate for GAT layers",
        "gat_learning_rate":   "Learning rate for GAT optimizer",
        "gat_epochs":          "Number of training epochs for GAT",
        "gat_num_layers":      "Number of stacked GAT layers",
        "knn_k":               "Number of K-nearest neighbors for graph construction",
        "train_ratio":         "Percentage of data used for training",
        "test_ratio":          "Percentage of data used for testing",
        "data_split_ratio":    "Train/Test split ratio string (e.g., 80/20)",
    }

    headers1 = ["Parameter Name", "Value", "Description"]
    for ci, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=ci, value=h)
        header_style(cell)

    row = 2
    for k, v in settings.items():
        ws1.cell(row=row, column=1, value=k)
        ws1.cell(row=row, column=2, value=str(v))
        ws1.cell(row=row, column=3, value=PARAM_DESCRIPTIONS.get(k, "—"))
        row += 1
    auto_width(ws1)

    # ── SHEET 2: Training Metrics ─────────────────────────────────────
    ws2 = wb.create_sheet("Training Metrics")
    ws2.freeze_panes = "A2"
    metric_headers = ["Model", "Accuracy", "Precision", "Recall", "F1-Score",
                      "Macro Avg", "Weighted Avg", "MCC", "ROC-AUC", "Mean Std"]
    for ci, h in enumerate(metric_headers, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        header_style(cell, "#1A5E3A")

    ws2.cell(row=2, column=1,  value=r.model_name or "—")
    ws2.cell(row=2, column=2,  value=round(r.accuracy or 0, 4))
    ws2.cell(row=2, column=3,  value=round(r.precision or 0, 4))
    ws2.cell(row=2, column=4,  value=round(r.recall or 0, 4))
    ws2.cell(row=2, column=5,  value=round(r.f1_score or 0, 4))
    ws2.cell(row=2, column=6,  value=round(r.macro_average or 0, 4))
    ws2.cell(row=2, column=7,  value=round(r.weighted_average or 0, 4))
    ws2.cell(row=2, column=8,  value=round(r.mcc or 0, 4))
    ws2.cell(row=2, column=9,  value=round(r.roc_auc or 0, 4))
    ws2.cell(row=2, column=10, value=round(r.mean_std or 0, 4))
    auto_width(ws2)

    # ── SHEET 3: Testing Metrics ──────────────────────────────────────
    ws3 = wb.create_sheet("Testing Metrics")
    ws3.freeze_panes = "A2"
    for ci, h in enumerate(metric_headers, 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        header_style(cell, "#5E1A3A")

    if test_hist:
        ws3.cell(row=2, column=1,  value=test_hist.model_name or "—")
        ws3.cell(row=2, column=2,  value=round(test_hist.accuracy or 0, 4))
        ws3.cell(row=2, column=3,  value=round(test_hist.precision or 0, 4))
        ws3.cell(row=2, column=4,  value=round(test_hist.recall or 0, 4))
        ws3.cell(row=2, column=5,  value=round(test_hist.f1_score or 0, 4))
        ws3.cell(row=2, column=6,  value=round(test_hist.macro_average or 0, 4))
        ws3.cell(row=2, column=7,  value=round(test_hist.weighted_average or 0, 4))
        ws3.cell(row=2, column=8,  value=round(test_hist.mcc or 0, 4))
        ws3.cell(row=2, column=9,  value=round(test_hist.roc_auc or 0, 4))
        ws3.cell(row=2, column=10, value=round(test_hist.mean_std or 0, 4))
    else:
        ws3.cell(row=2, column=1, value="No testing data found — run a bulk prediction first")
    auto_width(ws3)

    # ── SHEET 4: Comparison ───────────────────────────────────────────
    ws4 = wb.create_sheet("Comparison")
    ws4.freeze_panes = "A2"
    comp_headers = ["Metric", "Training Score", "Testing Score", "Difference (%)", "Status"]
    for ci, h in enumerate(comp_headers, 1):
        cell = ws4.cell(row=1, column=ci, value=h)
        header_style(cell, "#4A3060")

    metrics_map = [
        ("Accuracy",         r.accuracy or 0,         test_hist.accuracy if test_hist else None),
        ("Precision",        r.precision or 0,        test_hist.precision if test_hist else None),
        ("Recall",           r.recall or 0,           test_hist.recall if test_hist else None),
        ("F1-Score",         r.f1_score or 0,         test_hist.f1_score if test_hist else None),
        ("Macro Average",    r.macro_average or 0,    test_hist.macro_average if test_hist else None),
        ("Weighted Average", r.weighted_average or 0, test_hist.weighted_average if test_hist else None),
        ("MCC",              r.mcc or 0,              test_hist.mcc if test_hist else None),
        ("ROC-AUC",          r.roc_auc or 0,          test_hist.roc_auc if test_hist else None),
        ("Mean Std",         r.mean_std or 0,         test_hist.mean_std if test_hist else None),
    ]

    for row_i, (metric, train_val, test_val) in enumerate(metrics_map, 2):
        ws4.cell(row=row_i, column=1, value=metric)
        ws4.cell(row=row_i, column=2, value=round(train_val, 4))
        if test_val is not None:
            diff_pct = abs(train_val - test_val) * 100
            status   = "OVERFIT" if diff_pct > 7.0 else "OK"
            ws4.cell(row=row_i, column=3, value=round(test_val, 4))
            ws4.cell(row=row_i, column=4, value=round(diff_pct, 2))
            status_cell = ws4.cell(row=row_i, column=5, value=status)
            if status == "OVERFIT":
                status_cell.font = Font(bold=True, color="C00000")
            else:
                status_cell.font = Font(bold=True, color="375623")
        else:
            ws4.cell(row=row_i, column=3, value="N/A")
            ws4.cell(row=row_i, column=4, value="N/A")
            ws4.cell(row=row_i, column=5, value="N/A")
    auto_width(ws4)

    # ── SHEET 5: Per-Epoch History ────────────────────────────────────
    ws5 = wb.create_sheet("Per-Epoch History")
    ws5.freeze_panes = "A2"
    epoch_headers = ["Fold", "Epoch", "Train Loss", "Val Loss",
                     "Accuracy", "Precision", "Recall", "F1", "MCC", "ROC-AUC",
                     "Train Acc", "Train F1", "Overfit Gap", "Is Best"]
    for ci, h in enumerate(epoch_headers, 1):
        cell = ws5.cell(row=1, column=ci, value=h)
        header_style(cell, "#2D4B73")

    for row_i, log in enumerate(epoch_logs, 2):
        ws5.cell(row=row_i, column=1,  value=log.get("fold", "—"))
        ws5.cell(row=row_i, column=2,  value=log.get("epoch", "—"))
        ws5.cell(row=row_i, column=3,  value=log.get("loss", 0))
        ws5.cell(row=row_i, column=4,  value=log.get("val_loss", 0))
        ws5.cell(row=row_i, column=5,  value=log.get("akurasi", 0))
        ws5.cell(row=row_i, column=6,  value=log.get("presisi", 0))
        ws5.cell(row=row_i, column=7,  value=log.get("recall", 0))
        ws5.cell(row=row_i, column=8,  value=log.get("f1", 0))
        ws5.cell(row=row_i, column=9,  value=log.get("mcc", 0))
        ws5.cell(row=row_i, column=10, value=log.get("roc_auc", 0))
        ws5.cell(row=row_i, column=11, value=log.get("train_akurasi", 0))
        ws5.cell(row=row_i, column=12, value=log.get("train_f1", 0))
        ws5.cell(row=row_i, column=13, value=log.get("overfit_gap", 0))
        ws5.cell(row=row_i, column=14, value="YES" if log.get("is_best") else "")
    auto_width(ws5)

    # ── Save to temp file ─────────────────────────────────────────────
    safe_name  = "".join(c if c.isalnum() or c in "-_" else "_" for c in (r.model_name or "model"))
    timestamp  = dt.now().strftime("%Y%m%d_%H%M%S")
    filename   = f"training_results_{safe_name}_{timestamp}.xlsx"
    tmp_path   = os.path.join(MODELS_DIR, filename)
    wb.save(tmp_path)

    return FileResponse(
        tmp_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )
